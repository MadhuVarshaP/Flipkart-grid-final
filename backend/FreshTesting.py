from flask import Flask, request, jsonify
from flask_cors import CORS
import cv2
from ultralytics import YOLO
import openpyxl
from datetime import datetime
import os

app = Flask(__name__)
CORS(app)

# Load the YOLO model
model_path = 'Freshnew100.pt'  # Path to your .pt model file
model = YOLO(model_path)  # Load the YOLO model

# Define label mapping (ensure it matches your YOLO training classes)
label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale', 
                 'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']

# Define expected lifespan for each product
expected_life_span = {
    "apple": 7,    # Apple lasts 7 days when fresh
    "onion": 10,   # Onion lasts 10 days when fresh
    "carrot": 5,   # Carrot lasts 5 days when fresh
    "tomato": 3    # Tomato lasts 3 days when fresh
}

# Confidence threshold for filtering low-confidence predictions
CONFIDENCE_THRESHOLD = 0.5

# Excel file for storing fresh count
excel_file = "detection_fresh_count3.xlsx"

# Initialize or load Excel workbook and sheet
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["S No", "Product", "Fresh Count", "Last Detected Time", "Expected Life Span"])

# Helper function to update fresh count in Excel
def update_fresh_count(product, is_fresh):
    lifespan = "N/A" if not is_fresh else expected_life_span.get(product, "Unknown")

    # Check if the product already exists in the sheet
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == product:  # Check the "Product" column
            product_found = True
            if is_fresh:  # Only increment count if the product is fresh
                row[2].value += 1  # Increment the fresh count
            row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Update the last detected time
            row[4].value = lifespan  # Update expected lifespan
            break

    if not product_found:
        # Add a new row for the product
        fresh_count = 1 if is_fresh else 0
        last_detected_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([sheet.max_row, product, fresh_count, last_detected_time, lifespan])

# Route for handling freshness detection
@app.route('/detect-freshness', methods=['POST'])
def detect_freshness():
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400

        image_file = request.files['image']
        uploads_dir = "uploads"
        os.makedirs(uploads_dir, exist_ok=True)  # Ensure the directory exists
        image_path = os.path.join(uploads_dir, image_file.filename)

        # Save the image temporarily
        image_file.save(image_path)

        # Process the image using YOLO
        image = cv2.imread(image_path)
        if image is None:
            return jsonify({"error": "Error reading image"}), 400

        results = model(image)
        detection_results = []

        for result in results:
            boxes = result.boxes.xyxy  # Bounding box coordinates
            confidences = result.boxes.conf  # Confidence scores
            labels = result.boxes.cls  # Class indices

            for i, box in enumerate(boxes):
                confidence = confidences[i].item()
                if confidence < CONFIDENCE_THRESHOLD:
                    continue  # Skip low-confidence predictions

                x1, y1, x2, y2 = map(int, box)
                label_idx = int(labels[i])
                predicted_label = label_encoder[label_idx]
                product, freshness = predicted_label.split('_')

                # Update fresh count in Excel
                is_fresh = (freshness == "fresh")
                update_fresh_count(product, is_fresh)

                # Append result for the frontend
                detection_results.append({
                    "product": product,
                    "freshness": freshness,
                    "confidence": confidence,
                    "bbox": [x1, y1, x2, y2]
                })

        # Save the workbook
        workbook.save(excel_file)

        return jsonify({
            "status": "Freshness detection completed",
            "detections": detection_results
        })

    except Exception as e:
        print(f"Error in detect_freshness route: {e}")
        return jsonify({"error": str(e)}), 500



if __name__ == '__main__':
    app.run(debug=True)





# import cv2
# from ultralytics import YOLO
# import openpyxl
# from datetime import datetime

# # Load the YOLO model
# model_path = 'Freshnew100.pt'  # Path to your .pt model file
# model = YOLO(model_path)  # Load the YOLO model

# # Define label mapping (ensure it matches your YOLO training classes)
# label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale', 
#                  'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']

# # Define expected lifespan for each product
# expected_life_span = {
#     "apple": 7,    # Apple lasts 7 days when fresh
#     "onion": 10,   # Onion lasts 10 days when fresh
#     "carrot": 5,   # Carrot lasts 5 days when fresh
#     "tomato": 3    # Tomato lasts 3 days when fresh
# }

# # Confidence threshold for filtering low-confidence predictions
# CONFIDENCE_THRESHOLD = 0.5

# # Excel file for storing fresh count
# excel_file = "detection_fresh_count3.xlsx"

# # Initialize or load Excel workbook and sheet
# try:
#     workbook = openpyxl.load_workbook(excel_file)
#     sheet = workbook.active
# except FileNotFoundError:
#     workbook = openpyxl.Workbook()
#     sheet = workbook.active
#     sheet.append(["S No", "Product", "Fresh Count", "Last Detected Time", "Expected Life Span"])

# # Helper function to update fresh count in Excel
# def update_fresh_count(product, is_fresh):
#     lifespan = "N/A" if not is_fresh else expected_life_span.get(product, "Unknown")

#     # Check if the product already exists in the sheet
#     product_found = False
#     for row in sheet.iter_rows(min_row=2, values_only=False):
#         if row[1].value == product:  # Check the "Product" column
#             product_found = True
#             if is_fresh:  # Only increment count if the product is fresh
#                 row[2].value += 1  # Increment the fresh count
#             row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Update the last detected time
#             row[4].value = lifespan  # Update expected lifespan
#             break

#     if not product_found:
#         # Add a new row for the product
#         fresh_count = 1 if is_fresh else 0
#         last_detected_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
#         sheet.append([sheet.max_row, product, fresh_count, last_detected_time, lifespan])

# # Detect and classify products in an image
# def detect_and_classify(image_path):
#     image = cv2.imread(image_path)
#     if image is None:
#         print("Error: Image not found.")
#         return None

#     # Use the YOLO model for prediction
#     results = model(image)
#     detected_results = []

#     # Iterate through detected objects
#     for result in results:
#         boxes = result.boxes.xyxy  # Bounding box coordinates
#         confidences = result.boxes.conf  # Confidence scores
#         labels = result.boxes.cls  # Class indices

#         for i, box in enumerate(boxes):
#             confidence = confidences[i].item()  # Confidence score
#             if confidence < CONFIDENCE_THRESHOLD:
#                 continue  # Skip low-confidence predictions

#             x1, y1, x2, y2 = map(int, box)  # Convert coordinates to integers
#             label_idx = int(labels[i])  # Class index
#             predicted_label = label_encoder[label_idx]  # Get label from mapping
#             product, freshness = predicted_label.split('_')  # Split into product and freshness

#             # Update fresh count in Excel
#             is_fresh = (freshness == "fresh")
#             update_fresh_count(product, is_fresh)

#             # Draw bounding box and label on the image
#             label_text = f"{product} ({freshness})"
#             cv2.rectangle(image, (x1 + 5, y1 + 5), (x2 - 5, y2 - 5), (0, 255, 0), 1)
#             cv2.putText(image, label_text, (x1, y1 - 10), cv2.FONT_HERSHEY_SIMPLEX, 0.4, (0, 255, 0), 1)

#     # Show the image with bounding boxes
#     cv2.imshow("Detection", image)
#     cv2.waitKey(0)
#     cv2.destroyAllWindows()

# # Main logic
# image_path = r'C:\Users\uvara\OneDrive\Desktop\lets begin\Flipkart product analysis\IMG20241012101311.jpg'
# detect_and_classify(image_path)

# # Save the workbook
# workbook.save(excel_file)
# print(f"Fresh count updated and saved to {excel_file}")







