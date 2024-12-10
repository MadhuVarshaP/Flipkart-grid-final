from flask import Flask, request, jsonify
from flask_cors import CORS
import tensorflow as tf
from werkzeug.utils import secure_filename
import cv2
import numpy as np
from ultralytics import YOLO
from tensorflow.keras.models import load_model
from tensorflow.keras.preprocessing.image import img_to_array
from tensorflow.keras.preprocessing import image
import easyocr
import json
import re
from datetime import datetime
import pandas as pd
import os
import openpyxl

app = Flask(__name__)
CORS(app)


# Load the saved model
model_path = 'my_model.keras'
model = tf.keras.models.load_model(model_path)
print("Model loaded successfully!")

# Load the class indices (mapping from class labels to indices)
class_indices_path = 'class_indices.json'
with open(class_indices_path, 'r') as f:
    class_indices = json.load(f)
label_map = {v: k for k, v in class_indices.items()}

# Ensure the uploads directory exists
os.makedirs("uploads", exist_ok=True)

# File path for the Excel sheet
excel_file_path = "product_predictions.xlsx"

# Create or load the Excel file
def initialize_excel():
    if not os.path.exists(excel_file_path):
        df = pd.DataFrame(columns=['S. No.', 'File Name', 'Predicted Brand', 'Timestamp'])
        df.to_excel(excel_file_path, index=False)

# Preprocess the image for prediction
def preprocess_image(img_path):
    img = image.load_img(img_path, target_size=(224, 224))  # Resize image to input size
    img_array = image.img_to_array(img)
    img_array = np.expand_dims(img_array, axis=0)  # Add batch dimension
    img_array /= 255.0  # Normalize pixel values
    return img_array

# Function to update Excel with new data
def update_excel(data):
    if os.path.exists(excel_file_path):
        df = pd.read_excel(excel_file_path)
    else:
        df = pd.DataFrame(columns=['S. No.', 'File Name', 'Predicted Brand', 'Timestamp'])

    # Add new entry to the DataFrame
    new_row = pd.DataFrame(data, columns=df.columns)
    df = pd.concat([df, new_row], ignore_index=True)
    df.to_excel(excel_file_path, index=False)

# Prediction API
@app.route('/predict-folder', methods=['POST'])
def predict_folder():
    # Check if files are present in the request
    if 'images' not in request.files:
        return jsonify({"error": "No files found in the request"}), 400

    images = request.files.getlist('images')
    
    # Check if any images were uploaded
    if not images or images[0].filename == '':
        return jsonify({"error": "No images uploaded"}), 400

    results = []
    excel_data = []

    # Loop through each file
    for idx, img_file in enumerate(images, start=1):
        try:
            # Secure and save the filename
            filename = secure_filename(img_file.filename)
            file_path = os.path.join('uploads', filename)
            img_file.save(file_path)

            # Preprocess the image
            img_array = preprocess_image(file_path)

            # Predict using the model
            predictions = model.predict(img_array)
            predicted_class = np.argmax(predictions[0])
            predicted_label = label_map.get(predicted_class, "Unknown")

            # Get current time
            current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

            # Prediction result
            result = {
                'filename': filename,
                'predicted_brand': predicted_label  # Key matches frontend expectation
            }
            results.append(result)

            # Prepare data for Excel
            excel_data.append([
                idx,  # S. No.
                filename,  # File Name
                predicted_label,  # Predicted Brand
                current_time  # Timestamp
            ])

            # Clean up the uploaded file
            os.remove(file_path)

        except Exception as e:
            results.append({
                'filename': filename,
                'error': str(e)
            })

    # Update Excel with prediction results
    if excel_data:
        update_excel(excel_data)

    return jsonify(results)  # Return full results list

# @app.route('/predict-folder', methods=['POST'])
# def predict_folder():
#     # Check if files are present in the request
#     if 'images' not in request.files:
#         return jsonify({"error": "No files found in the request"}), 400

#     images = request.files.getlist('images')
    
#     # Check if any images were uploaded
#     if not images or images[0].filename == '':
#         return jsonify({"error": "No images uploaded"}), 400

#     results = []
#     excel_data = []

#     # Loop through each file
#     for idx, img_file in enumerate(images, start=1):
#         try:
#             # Secure and save the filename
#             filename = secure_filename(img_file.filename)
#             file_path = os.path.join('uploads', filename)
#             img_file.save(file_path)

#             # Preprocess the image
#             img_array = preprocess_image(file_path)

#             # Predict using the model
#             predictions = model.predict(img_array)
#             predicted_class = np.argmax(predictions[0])
#             predicted_label = label_map.get(predicted_class, "Unknown")

#             # Get current time
#             current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

#             # Prediction result
#             result = {
#                 'filename': filename,
#                 'predicted_brand': predicted_label
#             }
#             results.append(result)

#             # Prepare data for Excel
#             excel_data.append([
#                 idx,  # S. No.
#                 filename,  # File Name
#                 predicted_label,  # Predicted Brand
#                 current_time  # Timestamp
#             ])

#             # Clean up the uploaded file
#             os.remove(file_path)

#         except Exception as e:
#             results.append({
#                 'filename': filename,
#                 'error': str(e)
#             })

#     # Update Excel with prediction results
#     if excel_data:
#         update_excel(excel_data)

#     return jsonify(results)

# Load models and resources
yolo_model = YOLO('best1143images(100).pt')  # YOLO model for expiry detection
expiry_keras_model = load_model('date_detection_model1.keras')  # Keras model for expiry detection
brand_keras_model = load_model('my_model.keras')  # Keras model for brand detection

# Load class indices for brand detection
with open('class_indices.json', 'r') as f:
    class_indices = json.load(f)

# Reverse the class indices mapping for decoding
class_indices_reversed = {v: k for k, v in class_indices.items()}

# Initialize EasyOCR reader
reader = easyocr.Reader(['en'])

# Function to preprocess the image for Keras model
def preprocess_for_keras(image):
    image = cv2.resize(image, (224, 224))  # Resize to MobileNetV2 input size
    image = image.astype("float32") / 255.0  # Normalize pixel values
    image = img_to_array(image)  # Convert to array
    image = np.expand_dims(image, axis=0)  
    return image

# Function to detect the brand
def detect_brand(image_path):
    image = cv2.imread(image_path)
    preprocessed_image = preprocess_for_keras(image)
    predictions = brand_keras_model.predict(preprocessed_image)
    predicted_class = np.argmax(predictions, axis=1)[0]
    brand_name = class_indices_reversed.get(predicted_class, "Unknown")
    return brand_name

# Function to extract dates from text using regex
def extract_dates(text):
    date_patterns = [
        r'\b\d{2}/\d{2}/\d{4}\b',    # MM/DD/YYYY or DD/MM/YYYY
        r'\b\d{4}/\d{2}/\d{2}\b',    # YYYY/MM/DD
        r'\b\d{2}\.\d{2}\.\d{4}\b',  # MM.DD.YYYY or DD.MM.YYYY
        r'\b\d{4}\.\d{2}\.\d{2}\b',  # YYYY.MM.DD
        r'\b\d{2}-\d{2}-\d{4}\b',    # MM-DD-YYYY or DD-MM-YYYY
        r'\b\d{4}-\d{2}-\d{2}\b',    # YYYY-MM-DD
    ]
    detected_dates = []
    for pattern in date_patterns:
        matches = re.findall(pattern, text)
        for match in matches:
            try:
                detected_dates.append(parse_date(match))
            except ValueError:
                continue
    return detected_dates

# Function to parse date string to datetime object
def parse_date(date_str):
    date_formats = [
        "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d",
        "%m.%d.%Y", "%d.%m.%Y", "%Y.%m.%d",
        "%m-%d-%Y", "%d-%m-%Y", "%Y-%m-%d",
    ]
    for date_format in date_formats:
        try:
            return datetime.strptime(date_str, date_format)
        except ValueError:
            continue
    raise ValueError(f"Date format not recognized: {date_str}")

# Function to select the latest date
def get_latest_date(dates):
    if not dates:
        return None
    return max(dates)

# Function to compare dates and determine expiration status
def compare_dates(latest_date):
    current_date = datetime.now()

    if latest_date < current_date:
        expired_status = "Yes"
        status_message = f"Expired on {latest_date.strftime('%d-%m-%Y')}"
        life_span = "NA"
    else:
        expired_status = "No"
        status_message = f"Valid until {latest_date.strftime('%d-%m-%Y')}"
        life_span = (latest_date - current_date).days

    return expired_status, life_span, status_message

# Function to save details to the Excel file
def save_details_to_excel(time, brand_name, expiry_date, expired_status, life_span, output_file):
    if os.path.exists(output_file):
        df = pd.read_excel(output_file)
    else:
        df = pd.DataFrame(columns=['Sno', 'Time', 'BrandName', 'Expiry Date', 'Expired Status', 'Life Span'])

    # Add new entry to the DataFrame
    new_row = pd.DataFrame({
        'Sno': [len(df) + 1],
        'Time': [time],
        'BrandName': [brand_name],
        'Expiry Date': [expiry_date.strftime('%d-%m-%Y') if expiry_date else "NA"],
        'Expired Status': [expired_status],
        'Life Span': [life_span if life_span != "NA" else "NA"]
    })
    df = pd.concat([df, new_row], ignore_index=True)

    # Save to the Excel file
    df.to_excel(output_file, index=False)
    print(f"Details saved to {output_file}")

@app.route('/recognize-expiry-date', methods=['POST'])
def recognize_expiry_date():
    if 'front_image' not in request.files or 'back_image' not in request.files:
        return jsonify({"error": "Both front and back images are required"}), 400

    front_image = request.files['front_image']
    back_image = request.files['back_image']

    # Save temporary images
    front_path = "temp_front.jpg"
    back_path = "temp_back.jpg"
    front_image.save(front_path)
    back_image.save(back_path)

    output_excel_file = 'Expiry_Brand_Details3.xlsx'

    try:
        # Call the process_images_and_save_details function to process the images
        brand_name = detect_brand(front_path)
        expiry_image = cv2.imread(back_path)
        results = yolo_model(expiry_image)

        detected_dates = []
        for result in results:
            boxes = result.boxes.xyxy
            confidences = result.boxes.conf

            for i, box in enumerate(boxes):
                x1, y1, x2, y2 = map(int, box)
                detected_region = expiry_image[y1:y2, x1:x2]
                preprocessed_region = preprocess_for_keras(detected_region)
                prediction = expiry_keras_model.predict(preprocessed_region)[0][0]

                if prediction < 0.5:
                    ocr_results = reader.readtext(detected_region)
                    for _, text, _ in ocr_results:
                        detected_dates.extend(extract_dates(text))

        # Determine the latest date
        latest_date = get_latest_date(detected_dates)
        expired_status, life_span, status_message = compare_dates(latest_date)

        # Save to Excel
        current_time = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        save_details_to_excel(current_time, brand_name, latest_date, expired_status, life_span, output_excel_file)

        # Construct response
        response = {
            "brand_name": brand_name,
            "expiry_date": latest_date.strftime('%d-%m-%Y') if latest_date else "Not Detected",
            "expired_status": expired_status,
            "life_span": life_span if life_span != "NA" else "Not Applicable",
            "status_message": status_message
        }

        return jsonify(response)

    except Exception as e:
        return jsonify({"error": str(e)}), 500

    finally:
        # Cleanup temporary files
        os.remove(front_path)
        os.remove(back_path)


# Load the YOLO model
model_path = 'Freshnew100.pt'  # Path to your .pt model file
model = YOLO(model_path)  # Load the YOLO model

# Define label mapping (ensure it matches your YOLO training classes)
label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale', 
                 'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']

# Define expected lifespan for each product
expected_life_span = {
    "apple": 7,    # Apple lasts 7 days when fresh
    "onion": 10,   # Onion lasts 10 days when fresh
    "carrot": 5,   # Carrot lasts 5 days when fresh
    "tomato": 3    # Tomato lasts 3 days when fresh
}

# Confidence threshold for filtering low-confidence predictions
CONFIDENCE_THRESHOLD = 0.5

# Excel file for storing fresh count
excel_file = "detection_fresh_count3.xlsx"

# Initialize or load Excel workbook and sheet
try:
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(["S No", "Product", "Fresh Count", "Last Detected Time", "Expected Life Span"])

# Helper function to update fresh count in Excel
def update_fresh_count(product, is_fresh):
    lifespan = "N/A" if not is_fresh else expected_life_span.get(product, "Unknown")

    # Check if the product already exists in the sheet
    product_found = False
    for row in sheet.iter_rows(min_row=2, values_only=False):
        if row[1].value == product:  # Check the "Product" column
            product_found = True
            if is_fresh:  # Only increment count if the product is fresh
                row[2].value += 1  # Increment the fresh count
            row[3].value = datetime.now().strftime("%Y-%m-%d %H:%M:%S")  # Update the last detected time
            row[4].value = lifespan  # Update expected lifespan
            break

    if not product_found:
        # Add a new row for the product
        fresh_count = 1 if is_fresh else 0
        last_detected_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        sheet.append([sheet.max_row, product, fresh_count, last_detected_time, lifespan])

# Route for handling freshness detection
@app.route('/detect-freshness', methods=['POST'])
def detect_freshness():
    try:
        if 'image' not in request.files:
            return jsonify({"error": "No image file provided"}), 400

        image_file = request.files['image']
        uploads_dir = "uploads"
        os.makedirs(uploads_dir, exist_ok=True)  # Ensure the directory exists
        image_path = os.path.join(uploads_dir, image_file.filename)

        # Save the image temporarily
        image_file.save(image_path)

        # Process the image using YOLO
        image = cv2.imread(image_path)
        if image is None:
            return jsonify({"error": "Error reading image"}), 400

        results = model(image)
        detection_results = []

        for result in results:
            boxes = result.boxes.xyxy  # Bounding box coordinates
            confidences = result.boxes.conf  # Confidence scores
            labels = result.boxes.cls  # Class indices

            for i, box in enumerate(boxes):
                confidence = confidences[i].item()
                if confidence < CONFIDENCE_THRESHOLD:
                    continue  # Skip low-confidence predictions

                x1, y1, x2, y2 = map(int, box)
                label_idx = int(labels[i])
                predicted_label = label_encoder[label_idx]
                product, freshness = predicted_label.split('_')

                # Update fresh count in Excel
                is_fresh = (freshness == "fresh")
                update_fresh_count(product, is_fresh)

                # Append result for the frontend
                detection_results.append({
                    "product": product,
                    "freshness": freshness,
                    "confidence": confidence,
                    "bbox": [x1, y1, x2, y2]
                })

        # Save the workbook
        workbook.save(excel_file)

        return jsonify({
            "status": "Freshness detection completed",
            "detections": detection_results
        })

    except Exception as e:
        print(f"Error in detect_freshness route: {e}")
        return jsonify({"error": str(e)}), 500



if __name__ == '__main__':
    initialize_excel()
    app.run(debug=True)















# from flask import Flask, request, jsonify
# from flask_cors import CORS
# import cv2
# import numpy as np
# from ultralytics import YOLO
# from tensorflow.keras.models import load_model
# from tensorflow.keras.preprocessing.image import img_to_array
# import easyocr
# import json
# import re
# from datetime import datetime
# import pandas as pd
# import os
# import tensorflow as tf
# from werkzeug.utils import secure_filename
# import openpyxl

# app = Flask(__name__)
# CORS(app)

# # Directories
# UPLOADS_DIR = "uploads"
# os.makedirs(UPLOADS_DIR, exist_ok=True)

# # Load Models
# yolo_model = YOLO('best1143images(100).pt')
# expiry_keras_model = load_model('date_detection_model1.keras')
# brand_keras_model = load_model('my_model.keras')
# yolo_freshness_model = YOLO('Freshnew100.pt')

# # Load class indices
# with open('class_indices.json', 'r') as f:
#     class_indices = json.load(f)
# class_indices_reversed = {v: k for k, v in class_indices.items()}

# # EasyOCR Reader
# reader = easyocr.Reader(['en'])

# # Freshness Data
# label_encoder = ['apple_fresh', 'apple_stale', 'onion_fresh', 'onion_stale', 
#                  'carrot_fresh', 'carrot_stale', 'tomato_fresh', 'tomato_stale']
# expected_life_span = {"apple": 7, "onion": 10, "carrot": 5, "tomato": 3}

# CONFIDENCE_THRESHOLD = 0.5

# # Utility Functions
# def preprocess_for_keras(image):
#     image = cv2.resize(image, (224, 224))
#     image = image.astype("float32") / 255.0
#     image = img_to_array(image)
#     return np.expand_dims(image, axis=0)

# def preprocess_image(img_path):
#     img = tf.keras.preprocessing.image.load_img(img_path, target_size=(224, 224))
#     img_array = tf.keras.preprocessing.image.img_to_array(img)
#     return np.expand_dims(img_array / 255.0, axis=0)

# def extract_dates(text):
#     date_patterns = [
#         r'\b\d{2}/\d{2}/\d{4}\b', r'\b\d{4}/\d{2}/\d{2}\b',
#         r'\b\d{2}\.\d{2}\.\d{4}\b', r'\b\d{4}\.\d{2}\.\d{2}\b',
#         r'\b\d{2}-\d{2}-\d{4}\b', r'\b\d{4}-\d{2}-\d{2}\b'
#     ]
#     detected_dates = []
#     for pattern in date_patterns:
#         matches = re.findall(pattern, text)
#         for match in matches:
#             try:
#                 detected_dates.append(datetime.strptime(match, "%d-%m-%Y"))
#             except ValueError:
#                 continue
#     return detected_dates

# def get_latest_date(dates):
#     return max(dates) if dates else None

# def save_to_excel(data, excel_file, columns):
#     if os.path.exists(excel_file):
#         df = pd.read_excel(excel_file)
#     else:
#         df = pd.DataFrame(columns=columns)
#     new_row = pd.DataFrame([data], columns=columns)
#     df = pd.concat([df, new_row], ignore_index=True)
#     df.to_excel(excel_file, index=False)

# def initialize_workbook(excel_file, headers):
#     try:
#         workbook = openpyxl.load_workbook(excel_file)
#         sheet = workbook.active
#     except FileNotFoundError:
#         workbook = openpyxl.Workbook()
#         sheet = workbook.active
#         sheet.append(headers)
#     return workbook, sheet

# # Routes
# @app.route('/recognize-expiry-date', methods=['POST'])
# def recognize_expiry_date():
#     try:
#         front_image = request.files.get('front_image')
#         back_image = request.files.get('back_image')
#         if not front_image or not back_image:
#             return jsonify({"error": "Both front and back images are required"}), 400

#         front_path = os.path.join(UPLOADS_DIR, "front_temp.jpg")
#         back_path = os.path.join(UPLOADS_DIR, "back_temp.jpg")
#         front_image.save(front_path)
#         back_image.save(back_path)

#         brand_name = detect_brand(front_path)
#         expiry_image = cv2.imread(back_path)
#         results = yolo_model(expiry_image)

#         detected_dates = []
#         for result in results:
#             for box in result.boxes.xyxy:
#                 x1, y1, x2, y2 = map(int, box)
#                 detected_region = expiry_image[y1:y2, x1:x2]
#                 preprocessed_region = preprocess_for_keras(detected_region)
#                 prediction = expiry_keras_model.predict(preprocessed_region)[0][0]
#                 if prediction < 0.5:
#                     ocr_results = reader.readtext(detected_region)
#                     for _, text, _ in ocr_results:
#                         detected_dates.extend(extract_dates(text))

#         latest_date = get_latest_date(detected_dates)
#         expired_status = "Yes" if latest_date and latest_date < datetime.now() else "No"
#         life_span = (latest_date - datetime.now()).days if latest_date else "N/A"

#         data = {
#             "Brand Name": brand_name,
#             "Expiry Date": latest_date.strftime('%d-%m-%Y') if latest_date else "Not Found",
#             "Expired": expired_status,
#             "Life Span": life_span
#         }

#         save_to_excel(data, "Expiry_Details.xlsx", ["Brand Name", "Expiry Date", "Expired", "Life Span"])
#         return jsonify(data)
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @app.route('/predict-folder', methods=['POST'])
# def predict_folder():
#     try:
#         images = request.files.getlist('images')
#         if not images:
#             return jsonify({"error": "No images provided"}), 400

#         results = []
#         for image in images:
#             image_path = os.path.join(UPLOADS_DIR, secure_filename(image.filename))
#             image.save(image_path)
#             img_array = preprocess_image(image_path)
#             predictions = brand_keras_model.predict(img_array)
#             predicted_class = np.argmax(predictions[0])
#             results.append(class_indices_reversed.get(predicted_class, "Unknown"))
#             os.remove(image_path)

#         return jsonify({"results": results})
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# @app.route('/detect-freshness', methods=['POST'])
# def detect_freshness():
#     try:
#         image_file = request.files.get('image')
#         if not image_file:
#             return jsonify({"error": "No image provided"}), 400

#         image_path = os.path.join(UPLOADS_DIR, secure_filename(image_file.filename))
#         image_file.save(image_path)
#         image = cv2.imread(image_path)

#         results = yolo_freshness_model(image)
#         detections = []
#         for result in results:
#             for i, box in enumerate(result.boxes.xyxy):
#                 confidence = result.boxes.conf[i]
#                 if confidence > CONFIDENCE_THRESHOLD:
#                     x1, y1, x2, y2 = map(int, box)
#                     label_idx = int(result.boxes.cls[i])
#                     label = label_encoder[label_idx]
#                     detections.append({"label": label, "confidence": confidence, "bbox": [x1, y1, x2, y2]})

#         return jsonify({"detections": detections})
#     except Exception as e:
#         return jsonify({"error": str(e)}), 500

# if __name__ == '__main__':
#     app.run(debug=True)
